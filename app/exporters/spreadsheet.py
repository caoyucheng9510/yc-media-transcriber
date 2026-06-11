from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.exporters.formats import document_polished_text
from app.schemas import JobRecord


XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXCEL_CELL_LIMIT = 32767
POLISHED_TEXT_CHUNK_SIZE = 30000

_TRUNCATED_MARKER = "[truncated]"
_ILLEGAL_EXCEL_CHARACTERS = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")


@dataclass(frozen=True)
class SpreadsheetRow:
    link: str
    title: str
    summary: str
    polished_chunks: list[str]


def build_batch_xlsx(jobs: list[JobRecord]) -> BytesIO:
    rows = [_job_to_spreadsheet_row(job) for job in jobs]
    max_polished_chunks = max((len(row.polished_chunks) for row in rows), default=1)
    headers = ["链接", "标题", "总结", "校对稿"]
    headers.extend(f"校对稿_{index}" for index in range(2, max_polished_chunks + 1))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "内容"
    sheet.append(headers)

    for row in rows:
        values = [row.link, row.title, row.summary]
        values.extend(row.polished_chunks)
        values.extend("" for _ in range(max_polished_chunks - len(row.polished_chunks)))
        sheet.append(values)

    _style_sheet(sheet, len(headers))

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def sanitize_cell_text(value: Any, *, truncate: bool = True) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\t", "")
    text = _ILLEGAL_EXCEL_CHARACTERS.sub("", text)
    if text.startswith(_FORMULA_PREFIXES):
        text = f"'{text}"
    if truncate:
        text = _truncate_cell_text(text)
    return text


def _job_to_spreadsheet_row(job: JobRecord) -> SpreadsheetRow:
    result = job.result or {}
    metadata = result.get("metadata") if isinstance(result.get("metadata"), dict) else {}
    title = job.title or metadata.get("title") or job.id
    polished_text = document_polished_text(result)
    return SpreadsheetRow(
        link=sanitize_cell_text(_job_link(job, metadata)),
        title=sanitize_cell_text(title),
        summary=sanitize_cell_text(result.get("summary")),
        polished_chunks=_split_polished_text(polished_text),
    )


def _job_link(job: JobRecord, metadata: dict[str, Any]) -> str:
    creator_import = metadata.get("creator_import")
    if isinstance(creator_import, dict):
        source_url = creator_import.get("source_url")
        if source_url:
            return str(source_url)

    for key in ("source_url", "display_url"):
        value = metadata.get(key)
        if value:
            return str(value)

    if job.source_type == "upload":
        return "本地上传"
    return job.source_value


def _split_polished_text(value: Any) -> list[str]:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\t", "")
    text = _ILLEGAL_EXCEL_CHARACTERS.sub("", text)
    if not text:
        return [""]
    return [
        sanitize_cell_text(text[index : index + POLISHED_TEXT_CHUNK_SIZE])
        for index in range(0, len(text), POLISHED_TEXT_CHUNK_SIZE)
    ]


def _truncate_cell_text(text: str) -> str:
    if len(text) <= EXCEL_CELL_LIMIT:
        return text
    keep = EXCEL_CELL_LIMIT - len(_TRUNCATED_MARKER)
    return f"{text[:keep]}{_TRUNCATED_MARKER}"


def _style_sheet(sheet: Any, column_count: int) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    header_font = Font(bold=True)
    alignment = Alignment(wrap_text=True, vertical="top")

    for cell in sheet[1]:
        cell.font = header_font

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    widths = [36, 28, 56]
    widths.extend(72 for _ in range(max(0, column_count - len(widths))))
    for index, width in enumerate(widths[:column_count], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
