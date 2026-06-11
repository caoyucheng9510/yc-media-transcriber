from __future__ import annotations

import json
import zipfile
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.config import Settings
from app.llm.processor import LLMProcessor
from app.main import create_app
from app.schemas import JobOptions, Segment
from app.storage import SQLiteStore
from app.terminology import TerminologyStore


def test_upload_job_completes_with_mock_asr(client: TestClient, sample_wav: Path) -> None:
    with sample_wav.open("rb") as fh:
        response = client.post(
            "/api/jobs/upload",
            files={"file": ("sample.wav", fh, "audio/wav")},
            data={"options": '{"llm_polish":false,"summary":false}'},
        )
    assert response.status_code == 200
    job_id = response.json()["job_id"]
    assert response.json()["view_url"] == "/"
    job = client.get(f"/api/jobs/{job_id}").json()
    assert job["status"] == "completed"
    result = client.get(f"/api/jobs/{job_id}/result").json()
    assert result["raw_transcript"] == "测试转录文本"
    assert result["structured_transcript"] == []
    assert result["llm_detail"]["enabled"] is False
    assert result["metadata"]["cache"]["hit"] is False
    assert "cache_key" in result["metadata"]["cache"]
    assert set(result["artifacts"]) == {"document_md", "document_pdf"}
    markdown = client.get(f"/api/jobs/{job_id}/artifacts/document_md")
    assert markdown.status_code == 200
    assert "sample.wav.md" in markdown.headers["content-disposition"]
    assert "## 总结" in markdown.text
    assert "## 校对稿" in markdown.text
    assert "测试转录文本" in markdown.text
    pdf = client.get(f"/api/jobs/{job_id}/artifacts/document_pdf")
    assert pdf.status_code == 200
    assert pdf.headers["content-type"] == "application/pdf"
    assert pdf.content.startswith(b"%PDF")
    unsupported_artifact = client.get(f"/api/jobs/{job_id}/artifacts/unsupported")
    assert unsupported_artifact.status_code == 404


def test_api_result_and_markdown_use_merged_public_polished_text(tmp_path: Path, sample_wav: Path) -> None:
    class FakeTranscriber:
        def transcribe(self, audio_path: Path, options: JobOptions) -> list[Segment]:
            return [
                Segment(start=0.0, end=0.5, speaker="Speaker1", text="你好"),
                Segment(start=0.6, end=1.0, speaker="Speaker1", text="继续"),
            ]

    class DialogClient:
        available = True
        provider_name = "fake"
        missing_configuration_reason = None

        def chat(self, *, model: str, messages: list[dict[str, str]], timeout: float = 60.0) -> str:
            prompt = messages[-1]["content"]
            if "names、places" in prompt:
                return "{}"
            if "speaker_mapping" in prompt:
                return '{"speaker_mapping":{"Speaker1":"张三"},"confidence":{"Speaker1":0.9},"source_labels":["Speaker1"]}'
            if "calibrated_dialogs" in prompt:
                return (
                    '{"calibrated_dialogs":['
                    '{"speaker_label":"Speaker1","text":"你好。"},'
                    '{"speaker_label":"Speaker1","text":"继续。"}'
                    ']}'
                )
            if "固定输出 Markdown" in prompt:
                return "## 总结\n总结\n\n## 关键要点\n- A"
            return "unexpected"

    data_dir = tmp_path / "data"
    settings = Settings(
        app_data_dir=data_dir,
        asr_engine="mock",
        asr_model_dir=data_dir / "models",
        terms_path=data_dir / "terms.json",
        app_process_jobs_inline=True,
        llm_api_key="dummy",
        llm_summary_min_chars=0,
        llm_quality_min_ratio=0.0,
        llm_quality_max_ratio=100.0,
    )
    app = create_app(settings)
    with TestClient(app) as test_client:
        test_client.app.state.queue.processor.transcriber = FakeTranscriber()
        test_client.app.state.queue.processor.llm = LLMProcessor(
            settings,
            TerminologyStore(settings.terms_path),
            client=DialogClient(),
        )

        with sample_wav.open("rb") as fh:
            response = test_client.post(
                "/api/jobs/upload",
                files={"file": ("sample.wav", fh, "audio/wav")},
                data={"options": '{"speaker_diarization":true,"llm_polish":true,"summary":true}'},
            )

        assert response.status_code == 200
        job_id = response.json()["job_id"]
        result = test_client.get(f"/api/jobs/{job_id}/result").json()
        markdown = test_client.get(f"/api/jobs/{job_id}/artifacts/document_md")

    assert result["polished_text"] == "张三: 你好。继续。"
    assert len(result["structured_transcript"]) == 2
    assert result["structured_transcript"][0]["text"] == "你好。"
    assert result["structured_transcript"][1]["text"] == "继续。"
    assert markdown.status_code == 200
    assert "## 校对稿\n\n张三: 你好。继续。" in markdown.text


def test_capabilities_marks_missing_keys(client: TestClient) -> None:
    payload = client.get("/api/capabilities").json()
    assert payload["llm"]["available"] is False
    assert payload["platforms"]["douyin"]["available"] is False
    assert payload["inputs"]["creator_profile"]["available"] is False
    assert payload["exports"] == ["document_md", "document_pdf"]
    assert payload["batch_exports"] == ["document_md", "document_pdf", "spreadsheet_xlsx"]


def test_api_auth_token_required(tmp_path: Path) -> None:
    settings = Settings(
        app_data_dir=tmp_path / "data",
        asr_engine="mock",
        asr_model_dir=tmp_path / "data" / "models",
        terms_path=tmp_path / "data" / "terms.json",
        api_auth_token="token",
        app_process_jobs_inline=True,
    )
    app = create_app(settings)
    with TestClient(app) as authed:
        assert authed.get("/api/capabilities").status_code == 401
        assert authed.get("/api/capabilities", headers={"Authorization": "Bearer token"}).status_code == 200


def test_url_job_tikhub_missing_key_fails(client: TestClient, temp_settings: Settings) -> None:
    response = client.post(
        "/api/jobs",
        json={
            "source": {"type": "url", "value": "https://xhslink.com/example"},
            "options": {"llm_polish": False, "summary": False},
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]
    assert response.json()["view_url"] == "/"
    result = client.get(f"/api/jobs/{job_id}/result").json()
    assert result["status"] == "failed"
    assert result["error"]["code"] == "platform_provider_not_configured"
    assert not (temp_settings.temp_dir / job_id).exists()


def test_creator_preview_tikhub_missing_key_fails(client: TestClient) -> None:
    response = client.post(
        "/api/creator/preview",
        json={
            "input": "https://xhslink.com/m/example",
            "max_items": 20,
        },
    )

    assert response.status_code == 400
    assert response.json()["detail"]["code"] == "platform_provider_not_configured"


def test_creator_profile_input_is_rejected_before_job_creation(client: TestClient) -> None:
    response = client.post(
        "/api/jobs",
        json={
            "source": {"type": "url", "value": "https://xhslink.com/m/example"},
            "options": {"llm_polish": False, "summary": False},
        },
    )

    assert response.status_code == 400
    assert response.json()["detail"] == {
        "code": "creator_profile_input",
        "message": "这是创作者主页链接，请使用“创作者主页”入口先拉取作品清单，再选择要转录的视频。",
        "stage": "api",
    }
    assert client.get("/api/jobs").json()["items"] == []


def test_failed_job_can_be_retried_and_deleted(client: TestClient) -> None:
    response = client.post(
        "/api/jobs",
        json={
            "source": {"type": "url", "value": "https://xhslink.com/example"},
            "options": {"llm_polish": False, "summary": False},
        },
    )
    job_id = response.json()["job_id"]
    assert client.get(f"/api/jobs/{job_id}").json()["status"] == "failed"

    retry = client.post(f"/api/jobs/{job_id}/retry")

    assert retry.status_code == 200
    assert retry.json()["job_id"] == job_id
    assert retry.json()["view_url"] == "/"
    retried = client.get(f"/api/jobs/{job_id}").json()
    assert retried["status"] == "failed"
    assert retried["error"]["code"] == "platform_provider_not_configured"

    delete = client.delete(f"/api/jobs/{job_id}")

    assert delete.status_code == 204
    assert client.get(f"/api/jobs/{job_id}").status_code == 404


def test_batch_delete_jobs_deletes_terminal_and_skips_active(
    client: TestClient,
    temp_settings: Settings,
    sample_wav: Path,
) -> None:
    with sample_wav.open("rb") as fh:
        completed = client.post(
            "/api/jobs/upload",
            files={"file": ("completed.wav", fh, "audio/wav")},
            data={"options": '{"llm_polish":false,"summary":false}'},
        ).json()["job_id"]
    completed_result = client.get(f"/api/jobs/{completed}/result").json()
    completed_cache_key = completed_result["metadata"]["cache"]["cache_key"]
    store = SQLiteStore(temp_settings.db_path)
    completed_cache_record = store.get_media_cache_record(completed_cache_key)
    assert completed_cache_record is not None
    completed_cache_dir = Path(completed_cache_record["cache_dir"])
    assert completed_cache_dir.exists()
    failed = client.post(
        "/api/jobs",
        json={
            "source": {"type": "url", "value": "https://xhslink.com/example"},
            "options": {"llm_polish": False, "summary": False},
        },
    ).json()["job_id"]
    active = "job_active"
    store.create_job(
        active,
        "url",
        "https://example.com/a.mp4",
        {"llm_polish": False, "summary": False},
    )

    response = client.post(
        "/api/jobs/batch-delete",
        json={"job_ids": [completed, failed, active, "job_missing"]},
    )

    assert response.status_code == 200
    payload = response.json()
    assert set(payload["deleted"]) == {completed, failed}
    skipped = {item["job_id"]: item["code"] for item in payload["skipped"]}
    assert skipped == {
        active: "job_not_deletable",
        "job_missing": "job_not_found",
    }
    assert client.get(f"/api/jobs/{completed}").status_code == 404
    assert client.get(f"/api/jobs/{failed}").status_code == 404
    assert client.get(f"/api/jobs/{active}").status_code == 200
    assert store.get_media_cache_record(completed_cache_key) is None
    assert not completed_cache_dir.exists()


def test_batch_export_jobs_returns_markdown_zip_with_manifest(
    client: TestClient,
    temp_settings: Settings,
    sample_wav: Path,
) -> None:
    first = _upload_completed_job(client, sample_wav, "first.wav")
    second = _upload_completed_job(client, sample_wav, "second.wav")
    failed = client.post(
        "/api/jobs",
        json={
            "source": {"type": "url", "value": "https://xhslink.com/example"},
            "options": {"llm_polish": False, "summary": False},
        },
    ).json()["job_id"]
    missing_artifact = "job_no_artifact"
    store = SQLiteStore(temp_settings.db_path)
    store.create_job(
        missing_artifact,
        "url",
        "https://example.com/a.mp4",
        {"llm_polish": False, "summary": False},
    )
    store.update_job(missing_artifact, status="completed", progress=100, result={"artifacts": {}})

    response = client.post(
        "/api/jobs/batch-export",
        json={
            "job_ids": [first, second, failed, missing_artifact, "job_missing"],
            "artifact_type": "document_md",
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/zip"
    assert "transcripts-document-md" in response.headers["content-disposition"]
    with zipfile.ZipFile(BytesIO(response.content)) as archive:
        names = archive.namelist()
        assert "_manifest.json" in names
        markdown_names = [name for name in names if name.endswith(".md")]
        assert len(markdown_names) == 2
        assert "测试转录文本" in archive.read(markdown_names[0]).decode("utf-8")
        manifest = json.loads(archive.read("_manifest.json").decode("utf-8"))

    assert [item["job_id"] for item in manifest["exported"]] == [first, second]
    skipped = {item["job_id"]: item["code"] for item in manifest["skipped"]}
    assert skipped == {
        failed: "job_not_completed",
        missing_artifact: "artifact_not_found",
        "job_missing": "job_not_found",
    }


def test_batch_export_jobs_supports_pdf_zip(client: TestClient, sample_wav: Path) -> None:
    job_id = _upload_completed_job(client, sample_wav, "pdf.wav")

    response = client.post(
        "/api/jobs/batch-export",
        json={"job_ids": [job_id], "artifact_type": "document_pdf"},
    )

    assert response.status_code == 200
    assert "transcripts-document-pdf" in response.headers["content-disposition"]
    with zipfile.ZipFile(BytesIO(response.content)) as archive:
        pdf_names = [name for name in archive.namelist() if name.endswith(".pdf")]
        assert len(pdf_names) == 1
        assert archive.read(pdf_names[0]).startswith(b"%PDF")


def test_batch_export_jobs_returns_spreadsheet_xlsx_in_request_order(
    client: TestClient,
    temp_settings: Settings,
) -> None:
    store = SQLiteStore(temp_settings.db_path)
    first = _create_completed_result_job(
        store,
        "job_first",
        "url",
        "https://example.com/first-fallback",
        {
            "metadata": {
                "title": "First",
                "source_url": "https://example.com/first",
                "creator_import": {"source_url": "https://example.com/creator-first"},
            },
            "summary": "First summary",
            "polished_text": "First polished",
        },
    )
    second = _create_completed_result_job(
        store,
        "job_second",
        "upload",
        "/private/second.wav",
        {
            "metadata": {"title": "Second"},
            "summary": "Second summary",
            "polished_text": "Second polished",
        },
    )
    failed = _create_job(store, "job_failed", "url", "https://example.com/failed")
    store.update_job(failed, status="failed")
    queued = _create_job(store, "job_queued", "url", "https://example.com/queued")
    no_result = _create_job(store, "job_no_result", "url", "https://example.com/no-result")
    store.update_job(no_result, status="completed", progress=100)

    response = client.post(
        "/api/jobs/batch-export",
        json={
            "job_ids": [second, failed, first, "job_missing", second, queued, no_result],
            "artifact_type": "spreadsheet_xlsx",
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "transcripts-spreadsheet" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content))

    assert workbook.sheetnames == ["内容"]
    sheet = workbook["内容"]
    assert [cell.value for cell in sheet[1]] == ["链接", "标题", "总结", "校对稿"]
    assert sheet.max_row == 3
    assert [cell.value for cell in sheet[2]] == [
        "本地上传",
        "Second",
        "Second summary",
        "Second polished",
    ]
    assert [cell.value for cell in sheet[3]] == [
        "https://example.com/creator-first",
        "First",
        "First summary",
        "First polished",
    ]


def test_batch_export_spreadsheet_empty_returns_structured_error(
    client: TestClient,
    temp_settings: Settings,
) -> None:
    store = SQLiteStore(temp_settings.db_path)
    failed = _create_job(store, "job_failed", "url", "https://example.com/failed")
    store.update_job(failed, status="failed")
    no_result = _create_job(store, "job_no_result", "url", "https://example.com/no-result")
    store.update_job(no_result, status="completed", progress=100)

    response = client.post(
        "/api/jobs/batch-export",
        json={"job_ids": [failed, no_result, "job_missing"], "artifact_type": "spreadsheet_xlsx"},
    )

    assert response.status_code == 409
    assert response.json()["detail"]["code"] == "batch_export_empty"


def test_batch_export_empty_returns_structured_error(client: TestClient) -> None:
    failed = client.post(
        "/api/jobs",
        json={
            "source": {"type": "url", "value": "https://xhslink.com/example"},
            "options": {"llm_polish": False, "summary": False},
        },
    ).json()["job_id"]

    response = client.post(
        "/api/jobs/batch-export",
        json={"job_ids": [failed], "artifact_type": "document_md"},
    )

    assert response.status_code == 409
    assert response.json()["detail"]["code"] == "batch_export_empty"


def test_completed_job_is_not_retryable(client: TestClient, sample_wav: Path) -> None:
    with sample_wav.open("rb") as fh:
        response = client.post(
            "/api/jobs/upload",
            files={"file": ("sample.wav", fh, "audio/wav")},
            data={"options": '{"llm_polish":false,"summary":false}'},
        )
    job_id = response.json()["job_id"]

    retry = client.post(f"/api/jobs/{job_id}/retry")

    assert retry.status_code == 409
    assert retry.json()["detail"]["code"] == "job_not_retryable"


def test_api_errors_are_structured(client: TestClient) -> None:
    missing_job = client.get("/api/jobs/job_missing")
    assert missing_job.status_code == 404
    assert missing_job.json()["detail"]["code"] == "job_not_found"

    missing_artifact = client.get("/api/jobs/job_missing/artifacts/document_md")
    assert missing_artifact.status_code == 404
    assert missing_artifact.json()["detail"]["code"] == "artifact_not_found"


def test_upload_invalid_options_error_is_structured(client: TestClient, sample_wav: Path) -> None:
    with sample_wav.open("rb") as fh:
        response = client.post(
            "/api/jobs/upload",
            files={"file": ("sample.wav", fh, "audio/wav")},
            data={"options": "not-json"},
        )

    assert response.status_code == 400
    assert response.json()["detail"]["code"] == "invalid_options"


def test_upload_empty_file_error_is_structured(client: TestClient) -> None:
    response = client.post(
        "/api/jobs/upload",
        files={"file": ("empty.wav", BytesIO(b""), "audio/wav")},
        data={"options": '{"llm_polish":false,"summary":false}'},
    )

    assert response.status_code == 400
    assert response.json()["detail"]["code"] == "media_invalid"


def _upload_completed_job(client: TestClient, sample_wav: Path, filename: str) -> str:
    with sample_wav.open("rb") as fh:
        response = client.post(
            "/api/jobs/upload",
            files={"file": (filename, fh, "audio/wav")},
            data={"options": '{"llm_polish":false,"summary":false}'},
        )
    assert response.status_code == 200
    return response.json()["job_id"]


def _create_job(
    store: SQLiteStore,
    job_id: str,
    source_type: str,
    source_value: str,
) -> str:
    store.create_job(
        job_id,
        source_type,
        source_value,
        {"llm_polish": False, "summary": False},
    )
    return job_id


def _create_completed_result_job(
    store: SQLiteStore,
    job_id: str,
    source_type: str,
    source_value: str,
    result: dict,
) -> str:
    _create_job(store, job_id, source_type, source_value)
    store.update_job(job_id, status="completed", progress=100, result=result, title=result.get("metadata", {}).get("title"))
    return job_id
