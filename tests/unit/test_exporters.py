from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.exporters.formats import (
    document_polished_text,
    document_markdown,
    safe_export_stem,
    transcript_text,
    write_job_artifacts,
)
from app.exporters.spreadsheet import (
    EXCEL_CELL_LIMIT,
    POLISHED_TEXT_CHUNK_SIZE,
    build_batch_xlsx,
    sanitize_cell_text,
)
from app.schemas import JobRecord


def test_transcript_exports() -> None:
    segments = [{"start": 0, "end": 1.25, "speaker": "Speaker1", "text": "你好"}]
    assert transcript_text(segments) == "Speaker1: 你好"


def test_document_markdown_contains_summary_and_polished_text_only() -> None:
    markdown = document_markdown(
        {
            "job_id": "job_1",
            "metadata": {"title": "标题 A"},
            "summary": "## 总结\n\n摘要正文。\n\n## 关键要点\n\n- 要点一",
            "polished_text": "第一段校对稿。\n\n第二段校对稿。",
            "raw_transcript": "原始转录不应导出",
            "segments": [{"text": "segment 不应导出"}],
            "llm_detail": {"enabled": True, "prompt_version": "reference_style_v1"},
        }
    )

    assert markdown.startswith("# 标题 A\n\n## 总结")
    assert "## 校对稿" in markdown
    assert "第一段校对稿" in markdown
    assert "原始转录不应导出" not in markdown
    assert "segment 不应导出" not in markdown
    assert "llm_detail" not in markdown
    assert "reference_style_v1" not in markdown


def test_document_markdown_falls_back_to_structured_transcript_display_text() -> None:
    markdown = document_markdown(
        {
            "job_id": "job_1",
            "summary": None,
            "polished_text": "",
            "structured_transcript": [
                {
                    "start": 0.0,
                    "end": 0.5,
                    "speaker_label": "Speaker1",
                    "speaker_name": "张三",
                    "text": "你好。",
                },
                {
                    "start": 0.6,
                    "end": 1.0,
                    "speaker_label": "Speaker1",
                    "speaker_name": "张三",
                    "text": "继续。",
                },
            ],
        }
    )

    assert "## 校对稿\n\n张三: 你好。继续。" in markdown
    assert document_polished_text(
        {
            "polished_text": "",
            "structured_transcript": [
                {
                    "speaker_label": "Speaker1",
                    "speaker_name": "张三",
                    "text": "你好。",
                },
                {
                    "speaker_label": "Speaker1",
                    "speaker_name": "张三",
                    "text": "继续。",
                },
            ],
        }
    ) == "张三: 你好。继续。"


def test_write_job_artifacts_creates_markdown_and_pdf_with_title_filename(tmp_path: Path) -> None:
    artifacts = write_job_artifacts(
        tmp_path,
        {
            "job_id": "job_1",
            "metadata": {"title": "商业/内容: 方法论"},
            "summary": "## 总结\n\n摘要。\n\n## 关键要点\n\n- 要点",
            "polished_text": "整理后的文稿。",
        },
    )

    assert set(artifacts) == {"document_md", "document_pdf"}
    assert artifacts["document_md"].name == "商业内容 方法论.md"
    assert artifacts["document_pdf"].name == "商业内容 方法论.pdf"
    assert artifacts["document_md"].read_text(encoding="utf-8").startswith("# 商业/内容: 方法论")
    assert artifacts["document_pdf"].read_bytes().startswith(b"%PDF")


def test_safe_export_stem_falls_back_when_title_is_empty() -> None:
    assert safe_export_stem(" /:*? ", "job_1") == "job_1"


def test_batch_xlsx_contains_fixed_headers_and_url_priority() -> None:
    workbook = load_workbook(
        build_batch_xlsx(
            [
                _job(
                    "job_1",
                    source_value="https://example.com/fallback",
                    title="任务标题",
                    result={
                        "metadata": {
                            "title": "metadata 标题",
                            "source_url": "https://example.com/source",
                            "display_url": "https://example.com/display",
                            "creator_import": {"source_url": "https://example.com/creator"},
                        },
                        "summary": "摘要",
                        "polished_text": "校对稿",
                    },
                )
            ]
        )
    )

    sheet = workbook["内容"]
    assert [cell.value for cell in sheet[1]] == ["链接", "标题", "总结", "校对稿"]
    assert [cell.value for cell in sheet[2]] == [
        "https://example.com/creator",
        "任务标题",
        "摘要",
        "校对稿",
    ]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:D2"


def test_batch_xlsx_uses_local_upload_link_and_structured_transcript_fallback() -> None:
    workbook = load_workbook(
        build_batch_xlsx(
            [
                _job(
                    "job_upload",
                    source_type="upload",
                    source_value="/private/sample.wav",
                    result={
                        "metadata": {"title": "sample.wav"},
                        "summary": None,
                        "polished_text": "",
                        "structured_transcript": [
                            {
                                "speaker_label": "Speaker1",
                                "speaker_name": "张三",
                                "text": "你好。",
                            },
                            {
                                "speaker_label": "Speaker1",
                                "speaker_name": "张三",
                                "text": "继续。",
                            },
                        ],
                    },
                )
            ]
        )
    )

    sheet = workbook["内容"]
    assert sheet["A2"].value == "本地上传"
    assert sheet["B2"].value == "sample.wav"
    assert sheet["C2"].value is None
    assert sheet["D2"].value == "张三: 你好。继续。"


def test_batch_xlsx_splits_long_polished_text_into_continuation_columns() -> None:
    polished = "a" * POLISHED_TEXT_CHUNK_SIZE + "b" * POLISHED_TEXT_CHUNK_SIZE + "c"
    workbook = load_workbook(
        build_batch_xlsx(
            [
                _job(
                    "job_long",
                    result={
                        "metadata": {"title": "长文稿"},
                        "summary": "摘要",
                        "polished_text": polished,
                    },
                )
            ]
        )
    )

    sheet = workbook["内容"]
    assert [cell.value for cell in sheet[1]] == ["链接", "标题", "总结", "校对稿", "校对稿_2", "校对稿_3"]
    assert len(sheet["D2"].value) == POLISHED_TEXT_CHUNK_SIZE
    assert len(sheet["E2"].value) == POLISHED_TEXT_CHUNK_SIZE
    assert sheet["F2"].value == "c"


def test_batch_xlsx_protects_formula_like_cells() -> None:
    workbook = load_workbook(
        build_batch_xlsx(
            [
                _job(
                    "job_formula",
                    source_value="@link",
                    title="=title",
                    result={
                        "metadata": {},
                        "summary": "+summary",
                        "polished_text": "-polished",
                    },
                )
            ]
        )
    )

    sheet = workbook["内容"]
    assert sheet["A2"].value == "'@link"
    assert sheet["B2"].value == "'=title"
    assert sheet["C2"].value == "'+summary"
    assert sheet["D2"].value == "'-polished"


def test_sanitize_cell_text_normalizes_newlines_and_removes_invalid_characters() -> None:
    assert sanitize_cell_text("a\r\nb\rc\td\x00e\x0bf") == "a\nb\ncdef"


def test_sanitize_cell_text_truncates_non_polished_fields_with_marker() -> None:
    text = sanitize_cell_text("a" * (EXCEL_CELL_LIMIT + 10))

    assert len(text) == EXCEL_CELL_LIMIT
    assert text.endswith("[truncated]")


def _job(
    job_id: str,
    *,
    source_type: str = "url",
    source_value: str = "https://example.com/video",
    title: str | None = None,
    result: dict | None = None,
) -> JobRecord:
    return JobRecord(
        id=job_id,
        status="completed",
        source_type=source_type,
        source_value=source_value,
        options={},
        metadata={},
        result=result or {"metadata": {}, "summary": "", "polished_text": ""},
        progress=100,
        title=title,
        created_at="2026-06-10T00:00:00+00:00",
        updated_at="2026-06-10T00:00:00+00:00",
    )
